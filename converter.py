import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import CellCoordinatesException

PREOP_TARGET_COLUMN = "PreOp Target (Previous)"
POSTOP_TARGET_COLUMN = "PostOp Target (Current)"
STATUS_COLUMN = "PreOp or PostOp"
PATIENT_ID_COLUMN = "PatientID"

def load_mapping(mapping_path):
    """Load and validate the Excel mapping file."""

    try:
        mapping_df = pd.read_excel(
            mapping_path,
            dtype=str,
        )

    except FileNotFoundError as error:
        raise FileNotFoundError(
            f"The mapping Excel file could not be found:\n{mapping_path}"
        ) from error

    except Exception as error:
        raise ValueError(
            f"The mapping Excel file could not be opened:\n{error}"
        ) from error

    # Remove accidental spaces from column headings.
    mapping_df.columns = [
        str(column).strip()
        for column in mapping_df.columns
    ]

    required_columns = {
        "Export",
        PREOP_TARGET_COLUMN,
        POSTOP_TARGET_COLUMN,
    }

    missing_columns = required_columns - set(mapping_df.columns)

    if missing_columns:
        formatted_columns = "\n".join(
            f"• {column}"
            for column in sorted(missing_columns)
        )

        raise ValueError(
            "The mapping Excel file is missing the following "
            f"required column(s):\n\n{formatted_columns}"
        )

    # Remove completely empty spreadsheet rows.
    mapping_df = mapping_df.dropna(how="all")

    return mapping_df

def get_patient_ids(input_path):
    """Return the unique patient IDs contained in a REDCap export."""

    try:
        input_df = pd.read_csv(
            input_path,
            dtype=object,
        )

    except FileNotFoundError as error:
        raise FileNotFoundError(
            f"The input file could not be found:\n{input_path}"
        ) from error

    except Exception as error:
        raise ValueError(
            f"The REDCap export could not be opened:\n{error}"
        ) from error

    if input_df.empty:
        raise ValueError(
            "The selected REDCap export contains no records."
        )

    input_df.columns = [
        str(column).strip()
        for column in input_df.columns
    ]

    if PATIENT_ID_COLUMN not in input_df.columns:
        raise ValueError(
            "The selected export does not contain a "
            f"'{PATIENT_ID_COLUMN}' column."
        )

    patient_ids = []

    for value in input_df[PATIENT_ID_COLUMN]:
        if pd.isna(value):
            continue

        patient_id = str(value).strip()

        if not patient_id:
            continue

        if patient_id not in patient_ids:
            patient_ids.append(patient_id)

    if not patient_ids:
        raise ValueError(
            "No patient IDs were found in the selected export."
        )

    return patient_ids

def get_patient_data(input_df, patient_id):
    """
    Combine all non-blank values from rows belonging to one patient.

    Later non-blank values replace earlier non-blank values.
    """

    if PATIENT_ID_COLUMN not in input_df.columns:
        raise ValueError(
            "The selected export does not contain a "
            f"'{PATIENT_ID_COLUMN}' column."
        )

    patient_id_text = str(patient_id).strip()

    patient_rows = input_df[
        input_df[PATIENT_ID_COLUMN]
        .fillna("")
        .astype(str)
        .str.strip()
        == patient_id_text
    ]

    if patient_rows.empty:
        raise ValueError(
            f"Patient ID {patient_id_text} was not found "
            "in the selected export."
        )

    patient_data = {}

    for column in patient_rows.columns:
        non_blank_values = []

        for value in patient_rows[column]:
            if pd.isna(value):
                continue

            if isinstance(value, str) and value.strip() == "":
                continue

            non_blank_values.append(value)

        if non_blank_values:
            patient_data[column] = non_blank_values[-1]
        else:
            patient_data[column] = None

    return patient_data

def get_assessment_status(patient_data):
    """Read whether the selected patient is PreOp or PostOp."""

    if STATUS_COLUMN not in patient_data:
        raise ValueError(
            "The REDCap export does not contain the "
            f"'{STATUS_COLUMN}' column."
        )

    status = patient_data.get(STATUS_COLUMN)

    # Blank values default to PostOp.
    if pd.isna(status):
        return "postop"

    status_text = str(status).strip()

    if status_text == "":
        return "postop"

    normalised_status = (
        status_text
        .lower()
        .replace(" ", "")
        .replace("-", "")
    )

    if normalised_status == "preop":
        return "preop"

    if normalised_status == "postop":
        return "postop"

    raise ValueError(
        f"The '{STATUS_COLUMN}' field contains an "
        f"unrecognised value: {status!r}.\n\n"
        "Expected PreOp or PostOp."
    )

def is_blank_mapping_value(value):
    """
    Return True when a mapping cell should be ignored.

    This handles:
    - Empty Excel cells
    - Blank text
    - A dash used as a placeholder
    """

    if pd.isna(value):
        return True

    text = str(value).strip()

    return text in {"", "-"}

def get_usable_mapping_rows(mapping_df, target_column):
    """Return source-column and target-cell pairs for this assessment."""

    usable_rows = []

    for _, mapping_row in mapping_df.iterrows():
        source_value = mapping_row["Export"]
        target_value = mapping_row[target_column]

        if is_blank_mapping_value(source_value):
            continue

        if is_blank_mapping_value(target_value):
            continue

        source_column = str(source_value).strip()
        target_cell = str(target_value).strip()

        usable_rows.append(
            (source_column, target_cell)
        )

    return usable_rows

def get_patient_id_target(usable_mapping_rows):
    """Return Excel target cell mapped from PatientID"""

    for source_column, target_cell in usable_mapping_rows:
        if source_column == PATIENT_ID_COLUMN:
            return target_cell

    return None

def validate_columns(input_df, usable_mapping_rows):
    """Return mapped REDCap columns missing from the export."""

    required_columns = [
        source_column
        for source_column, _ in usable_mapping_rows
    ]

    missing_columns = [
        column
        for column in required_columns
        if column not in input_df.columns
    ]

    return sorted(set(missing_columns))

def is_invalid_merged_target(worksheet, target_cell):
    """
    Return True if the target is inside a merged range but is
    not the top-left cell of that range.
    """

    for merged_range in worksheet.merged_cells.ranges:
        if target_cell in merged_range:
            top_left_cell = worksheet.cell(
                row=merged_range.min_row,
                column=merged_range.min_col,
            ).coordinate

            return target_cell != top_left_cell

    return False

def validate_target_cells(worksheet, usable_mapping_rows):
    """Check that every mapping target is a valid Excel cell."""

    invalid_targets = []

    for source_column, target_cell in usable_mapping_rows:
        try:
            worksheet[target_cell]

            if is_invalid_merged_target(
                    worksheet,
                    target_cell,
            ):
                invalid_targets.append(
                    f"{source_column} → {target_cell} "
                    "(not the top-left cell of a merged range)"
                )

        except (ValueError, CellCoordinatesException):
            invalid_targets.append(
                f"{source_column} → {target_cell}"
            )

    if invalid_targets:
        formatted_targets = "\n".join(
            f"• {target}"
            for target in invalid_targets
        )

        raise ValueError(
            "The mapping file contains invalid target "
            f"cell references:\n\n{formatted_targets}"
        )

def cell_full(cell):
    """
    Returns True when Excel cell is not empty.
    """
    value = cell.value

    if value is None:
        return False

    if isinstance(value, str) and value.strip() == "":
        return False

    return True

def convert(
    input_path,
    mapping_path,
    template_path,
    output_path,
    patient_id,
    merge_path=None,
    progress_callback=None,
    cancel_flag=None,
):
    """Convert one REDCap CSV record into the Excel template."""

    # --------------------------------------------------
    # 1. Read the REDCap export
    # --------------------------------------------------

    try:
        input_df = pd.read_csv(input_path, dtype=object)

    except FileNotFoundError as error:
        raise FileNotFoundError(
            f"The input file could not be found:\n{input_path}"
        ) from error

    except Exception as error:
        raise ValueError(
            f"The REDCap export could not be opened:\n{error}"
        ) from error

    if input_df.empty:
        raise ValueError(
            "The selected REDCap export contains no records."
        )

    input_df.columns = [
        str(column).strip()
        for column in input_df.columns
    ]

    pt_data = get_patient_data(input_df, patient_id)

    if progress_callback:
        progress_callback(10)

    # --------------------------------------------------
    # 2. Determine PreOp or PostOp
    # --------------------------------------------------

    assessment_status = get_assessment_status(pt_data)

    if assessment_status == "preop":
        target_column = PREOP_TARGET_COLUMN
    else:
        target_column = POSTOP_TARGET_COLUMN

    if progress_callback:
        progress_callback(20)

    # --------------------------------------------------
    # 3. Load the mapping
    # --------------------------------------------------

    mapping_df = load_mapping(mapping_path)

    usable_mapping_rows = get_usable_mapping_rows(
        mapping_df,
        target_column,
    )

    if not usable_mapping_rows:
        raise ValueError(
            f"No {assessment_status.title()} targets were "
            "found in the mapping file."
        )

    missing_columns = validate_columns(
        input_df,
        usable_mapping_rows,
    )

    if missing_columns:
        formatted_columns = "\n".join(
            f"• {column}"
            for column in missing_columns
        )

        raise ValueError(
            "The REDCap export is missing the following "
            f"required column(s):\n\n{formatted_columns}"
        )

    if progress_callback:
        progress_callback(30)

    # --------------------------------------------------
    # 4. Open the physical-exam template
    # --------------------------------------------------

    workbook_source = merge_path or template_path

    try:
        workbook = load_workbook(workbook_source)

    except FileNotFoundError as error:
        if merge_path:
            file_description = "existing Physical Exam file"
        else:
            file_description = "Physical Exam template"

        raise FileNotFoundError(
            f"The {file_description} could not be found:\n{workbook_source}"
        ) from error

    except Exception as error:
        if merge_path:
            file_description = "existing Physical Exam file"
        else:
            file_description = "Physical Exam template"

        raise ValueError(
            f"The {file_description} could not be opened:\n{error}"
        ) from error

    worksheet = workbook.active

    validate_target_cells(worksheet, usable_mapping_rows)

    # Check patient ID before merging
    if merge_path:
        patient_id_target = get_patient_id_target(usable_mapping_rows)

        if patient_id_target is None:
            raise ValueError("The mapping file does not contain a PatientID target for this assessment type.\n\n"
                             "THe patient ID cannot be verified before merging.")

        existing_patient_id = worksheet[patient_id_target].value

        if pd.isna(existing_patient_id) or (isinstance(existing_patient_id, str) and existing_patient_id.strip() == ""):
            raise ValueError("The existing Physical Exam file does not contain a Patient ID.\n\n"
                             "THe file cannot be safely merged.")

        selected_patient_id = str(patient_id).strip()
        existing_patient_id = str(existing_patient_id).strip()

        if selected_patient_id != existing_patient_id:
            raise ValueError("The patient IDs do not match.\n\n"
                             f"Export patient ID: {selected_patient_id}\n"
                             f"Existing Physical Exam Patient ID: {existing_patient_id}\n\n"
                             "Please make sure both files belong to the same patient.")

    total_fields = len(usable_mapping_rows)
    written = 0
    skip_exist = 0
    skip_blank = 0

    # --------------------------------------------------
    # 5. Populate the template
    # --------------------------------------------------
    for index, (
        source_column,
        target_cell,
    ) in enumerate(usable_mapping_rows, start=1):

        if cancel_flag and cancel_flag():
            raise InterruptedError(
                "The conversion was cancelled."
            )

        value = pt_data[source_column]
        target = worksheet[target_cell]

        # Skip blank REDCap values.
        if pd.isna(value) or (
                isinstance(value, str)
                and value.strip() == ""
        ):
            skip_blank += 1

        # Protect any existing value in the workbook.
        elif cell_full(target):
            skip_exist += 1

        else:
            target.value = value
            written += 1

        if progress_callback:
            completed_fraction = index / total_fields
            progress = 30 + int(
                completed_fraction * 60
            )
            progress_callback(progress)

    # Check once more before saving.
    if cancel_flag and cancel_flag():
        raise InterruptedError(
            "The conversion was cancelled."
        )

    # --------------------------------------------------
    # 6. Save the completed workbook
    # --------------------------------------------------

    try:
        workbook.save(output_path)

    except PermissionError as error:
        raise PermissionError(
            "The output file could not be saved.\n\n"
            "Make sure the output file is not already open "
            "in Excel."
        ) from error

    except Exception as error:
        raise ValueError(
            f"The converted workbook could not be saved:\n{error}"
        ) from error

    if progress_callback:
        progress_callback(100)

    return {
        "output_path": output_path,
        "assessment_status": assessment_status,
        "merge_mode": merge_path is not None,
        "written_count": written,
        "skipped_existing_count": skip_exist,
        "skipped_blank_count": skip_blank,
    }